import openpyxl
book = openpyxl.load_workbook("C:\\Users\\suresha\\PycharmProjects\\Python_Selenium\\Udemy\\E2EFramework\\TestData\\demoData.xlsx")
sheet = book.active
list = []
dict = {}
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value = "Suresh"
print(sheet.max_column)
print(sheet.max_row)
print(sheet["D2"].value)
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(dict)