import openpyxl


class HomePageData:
    test_HomePage_data = [{"Name": "Suresh", "Password": "adef", "Email": "s@gmail.com", "Gender": "Male"},
                          {"Name": "Kumar", "Password": "das@31", "Email": "s@gmail.com", "Gender": "Female"}]

    @staticmethod
    def getData():
        book = openpyxl.load_workbook("C:\\Users\\suresha\\PycharmProjects\\Python_Selenium\\Udemy\\E2EFramework\\TestData\\demoData.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict]
